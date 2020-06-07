import MySQLdb
import openpyxl

wb=openpyxl.load_workbook('cotenbook.xlsx')
sheet = wb['sheet1']

vals = {}
idx = 4
cur_theme = ""
while (sheet.cell(idx, 2).value):
    temp = []
    for j in range(2,6):
        if sheet.cell(idx,j).value:
            temp.append(sheet.cell(idx,j).value)
        else:
            temp.append("")

    if cur_theme != temp[0]:
        cur_theme = temp[0]
        vals[cur_theme] = [[temp[2],temp[3]]]
    else:
        vals[cur_theme].append([temp[2],temp[3]])
    idx += 1

print(vals)
