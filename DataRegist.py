import MySQLdb
import openpyxl

# 接続する
conn = MySQLdb.connect(
    unix_socket = '/Applications/MAMP/tmp/mysql/mysql.sock',
    user='root',
    passwd='root',
    host='localhost',
    db='coten_listner')

# 接続を閉じる
conn.close

wb=openpyxl.load_workbook('koten_radio.xlsx')
sheet = wb['koten_radio']

vals = []
idx = 2
while (sheet.cell(idx, 1).value):
    temp = []
    for j in range(1,7):
        if sheet.cell(idx,j).value:
            temp.append(sheet.cell(idx,j).value)
        else:
            temp.append("")
    
    vals.append(temp)
    
    idx += 1

cursor = conn.cursor()
sql = "insert into tb_coten_radio (num, title, point, apple_podcast_url, google_podcast_url, youtube_url) values (%s,%s,%s,%s,%s,%s)"

for v in vals:
    cursor.execute(sql, v)

# 保存を実行（忘れると保存されないので注意）
conn.commit()
 
# 接続を閉じる
conn.close()